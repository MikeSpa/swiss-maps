#!/usr/bin/env python3
"""
Download and process Switzerland trade data for 2024.

Sources:
  BAZG (Federal Office for Customs and Border Security) - official 2024 actuals
  Exports: https://www.bazg.admin.ch/dam/en/sd-web/L0lqLFfjaSe2/2_4_LD_EXP_en.xlsx
  Imports: https://www.bazg.admin.ch/dam/en/sd-web/-bHd3OniokpL/2_4_LD_IMP_en.xlsx

  Both files contain two side-by-side ranking tables:
    Cols 0-5:  "Business cycle total" — excludes precious metals, gems, art, antiques
    Cols 6-11: "General total"        — includes everything

  We use the "Business cycle total" as primary metric (matches standard reporting of ~283B).

Output:
  public/trade/trade_2024.json
"""

import json
import sys
import requests
import openpyxl
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "public" / "trade"
OUT_FILE = OUT_DIR / "trade_2024.json"

BAZG_EXP_URL = "https://www.bazg.admin.ch/dam/en/sd-web/L0lqLFfjaSe2/2_4_LD_EXP_en.xlsx"
BAZG_IMP_URL = "https://www.bazg.admin.ch/dam/en/sd-web/-bHd3OniokpL/2_4_LD_IMP_en.xlsx"

# Minimum trade volume to include a country (CHF millions, either direction)
MIN_VOLUME_MN = 100

# BAZG country name → ISO-3166-1 alpha-2
NAME_TO_ISO2: dict[str, str] = {
    "USA":              "US",
    "Germany":          "DE",
    "France":           "FR",
    "Italy":            "IT",
    "United Kingdom":   "GB",
    "China":            "CN",
    "Japan":            "JP",
    "Spain":            "ES",
    "Austria":          "AT",
    "Netherlands":      "NL",
    "Belgium":          "BE",
    "Slovenia":         "SI",
    "Poland":           "PL",
    "Czech Republic":   "CZ",
    "Sweden":           "SE",
    "Denmark":          "DK",
    "Ireland":          "IE",
    "Portugal":         "PT",
    "Hungary":          "HU",
    "Romania":          "RO",
    "Slovakia":         "SK",
    "Finland":          "FI",
    "Bulgaria":         "BG",
    "Norway":           "NO",
    "Switzerland":      "CH",
    "Canada":           "CA",
    "Mexico":           "MX",
    "Brazil":           "BR",
    "Argentina":        "AR",
    "Australia":        "AU",
    "New Zealand":      "NZ",
    "India":            "IN",
    "Korea (South)":    "KR",
    "Singapore":        "SG",
    "Hong Kong":        "HK",
    "Taiwan":           "TW",
    "Thailand":         "TH",
    "Viet Nam":         "VN",
    "Malaysia":         "MY",
    "Indonesia":        "ID",
    "Bangladesh":       "BD",
    "Philippines":      "PH",
    "Pakistan":         "PK",
    "Emirates, Arab":   "AE",
    "Saudi Arabia":     "SA",
    "Qatar":            "QA",
    "Kuwait":           "KW",
    "Israel":           "IL",
    "Türkiye":          "TR",
    "Egypt":            "EG",
    "South Africa":     "ZA",
    "Nigeria":          "NG",
    "Ethiopia":         "ET",
    "Kenya":            "KE",
    "Morocco":          "MA",
    "Tunisia":          "TN",
    "Algeria":          "DZ",
    "Russia":           "RU",
    "Ukraine":          "UA",
    "Kazakhstan":       "KZ",
    "Serbia":           "RS",
    "Croatia":          "HR",
    "Greece":           "GR",
    "Lithuania":        "LT",
    "Latvia":           "LV",
    "Estonia":          "EE",
    "Belarus":          "BY",
    "Uzbekistan":       "UZ",
    "Azerbaijan":       "AZ",
    "Georgia":          "GE",
    "Armenia":          "AM",
    "Chile":            "CL",
    "Colombia":         "CO",
    "Peru":             "PE",
    "Ecuador":          "EC",
    "Venezuela":        "VE",
    "Cuba":             "CU",
    "Dominican Rep.":   "DO",
    "Costa Rica":       "CR",
    "Panama":           "PA",
    "Guatemala":        "GT",
    "Bolivia":          "BO",
    "Paraguay":         "PY",
    "Uruguay":          "UY",
    "Jordan":           "JO",
    "Lebanon":          "LB",
    "Bahrain":          "BH",
    "Oman":             "OM",
    "Iraq":             "IQ",
    "Iran":             "IR",
    "Libya":            "LY",
    "Ghana":            "GH",
    "Tanzania":         "TZ",
    "Ivory Coast":      "CI",
    "Angola":           "AO",
    "Mozambique":       "MZ",
    "Cameroon":         "CM",
    "Uganda":           "UG",
    "Zimbabwe":         "ZW",
    "Zambia":           "ZM",
    "Myanmar":          "MM",
    "Sri Lanka":        "LK",
    "Nepal":            "NP",
    "Cambodia":         "KH",
    "Laos":             "LA",
    "Mongolia":         "MN",
    "North Korea":      "KP",
    "Slovenia (2)":     "SI",  # fallback
    "Korea (North)":    "KP",
    "Moldova":          "MD",
    "Luxembourg":       "LU",
    "Malta":            "MT",
    "Cyprus":           "CY",
    "Iceland":          "IS",
    "Albania":          "AL",
    "Bosnia-Herzeg.":   "BA",
    "Kosovo":           "XK",
    "Montenegro":       "ME",
    "Macedonia, North": "MK",
    "North Macedonia":  "MK",
    "Liechtenstein":    "LI",
    "San Marino":       "SM",
}

# Geographic centroids [lng, lat] for map display
CENTROIDS: dict[str, list[float]] = {
    "US": [-98.5, 39.5],
    "DE": [10.5,  51.2],
    "FR": [2.3,   46.2],
    "IT": [12.6,  42.5],
    "GB": [-1.5,  52.5],
    "CN": [104.0, 35.0],
    "JP": [138.0, 36.0],
    "ES": [-3.7,  40.4],
    "AT": [14.6,  47.6],
    "NL": [5.3,   52.1],
    "BE": [4.5,   50.5],
    "SI": [14.8,  46.1],
    "PL": [19.1,  51.9],
    "CZ": [15.5,  49.8],
    "SE": [18.6,  60.1],
    "DK": [10.0,  56.3],
    "IE": [-8.2,  53.1],
    "PT": [-8.2,  39.5],
    "HU": [19.5,  47.2],
    "RO": [24.9,  45.9],
    "SK": [19.3,  48.7],
    "FI": [25.7,  64.0],
    "BG": [25.5,  42.7],
    "NO": [8.5,   60.5],
    "CA": [-96.8, 56.1],
    "MX": [-102.5,23.6],
    "BR": [-51.9, -14.2],
    "AR": [-63.6, -38.4],
    "AU": [133.8, -25.7],
    "NZ": [172.5, -40.9],
    "IN": [78.9,  20.6],
    "KR": [127.8, 35.9],
    "SG": [103.8, 1.3],
    "HK": [114.2, 22.3],
    "TW": [121.0, 23.7],
    "TH": [100.5, 15.9],
    "VN": [108.0, 14.1],
    "MY": [109.7, 4.2],
    "ID": [113.9, -0.8],
    "BD": [90.4,  23.7],
    "PH": [122.9, 12.9],
    "AE": [54.4,  23.4],
    "SA": [45.1,  24.0],
    "QA": [51.2,  25.3],
    "IL": [34.9,  31.5],
    "TR": [35.2,  38.9],
    "EG": [30.8,  26.8],
    "ZA": [25.1,  -29.0],
    "NG": [8.7,   9.1],
    "MA": [-7.1,  31.8],
    "RU": [97.7,  61.5],
    "UA": [32.0,  49.0],
    "RS": [21.0,  44.0],
    "HR": [15.2,  45.1],
    "GR": [21.8,  39.1],
    "CL": [-71.5, -35.7],
    "CO": [-74.3, 4.6],
    "PE": [-75.0, -9.2],
    "LU": [6.1,   49.8],
    "IS": [-18.5, 64.9],
    "KZ": [67.1,  48.0],
    "PK": [69.3,  30.4],
    "LK": [80.7,  7.9],
    # Previously missing — added to cover all 96 partners
    "KW": [47.5,  29.3],
    "LT": [23.9,  55.2],
    "TN": [9.6,   33.9],
    "CR": [-84.0,  9.7],
    "BH": [50.6,  26.0],
    "BA": [17.6,  44.2],
    "OM": [57.5,  21.5],
    "KH": [104.9, 12.6],
    "JO": [37.0,  31.2],
    "EE": [25.0,  58.6],
    "DZ": [3.0,   28.0],
    "LV": [24.6,  56.9],
    "PA": [-80.0,  8.5],
    "IR": [53.7,  32.4],
    "EC": [-78.1,  -1.8],
    "LB": [35.8,  33.9],
    "LY": [17.2,  26.3],
    "IQ": [44.4,  33.2],
    "MT": [14.4,  35.9],
    "UY": [-55.7, -32.5],
    "MK": [21.7,  41.6],
    "XK": [21.1,  42.6],
    "KE": [37.9,   0.0],
    "DO": [-70.2,  18.7],
    "GE": [43.4,  42.3],
    "GH": [-1.0,   7.9],
    "MM": [96.9,  16.9],
    "AZ": [47.6,  40.1],
    "UZ": [64.6,  41.3],
    "CY": [33.4,  35.1],
    "GT": [-90.2,  15.8],
    "AL": [20.2,  41.2],
    "BY": [27.9,  53.5],
}

# FTA status per ISO2 code
# Sources: SECO, EFTA FTA monitor
FTA_STATUS: dict[str, str] = {
    # EU / CH bilateral
    "DE": "EU_bilateral", "FR": "EU_bilateral", "IT": "EU_bilateral",
    "ES": "EU_bilateral", "NL": "EU_bilateral", "BE": "EU_bilateral",
    "AT": "EU_bilateral", "PL": "EU_bilateral", "SE": "EU_bilateral",
    "DK": "EU_bilateral", "FI": "EU_bilateral", "PT": "EU_bilateral",
    "IE": "EU_bilateral", "CZ": "EU_bilateral", "HU": "EU_bilateral",
    "RO": "EU_bilateral", "SK": "EU_bilateral", "BG": "EU_bilateral",
    "SI": "EU_bilateral", "HR": "EU_bilateral", "GR": "EU_bilateral",
    "LU": "EU_bilateral", "LT": "EU_bilateral", "LV": "EU_bilateral",
    "EE": "EU_bilateral", "MT": "EU_bilateral", "CY": "EU_bilateral",

    # EFTA / bilateral FTAs in force
    "CN": "in_force",
    "JP": "in_force",
    "CA": "in_force",
    "KR": "in_force",
    "MX": "in_force",
    "SG": "in_force",
    "HK": "in_force",
    "CL": "in_force",
    "CO": "in_force",
    "PE": "in_force",
    "MA": "in_force",
    "TN": "in_force",
    "EG": "in_force",
    "SA": "in_force",   # GCC FTA
    "AE": "in_force",   # GCC FTA
    "QA": "in_force",   # GCC FTA
    "KW": "in_force",   # GCC FTA
    "UA": "in_force",
    "ID": "in_force",
    "IS": "in_force",   # EFTA internal

    # Signed, not yet in force
    "IN": "signed_not_in_force",  # EFTA-India TEPA, signed 2024

    # Under negotiation
    "GB": "under_negotiation",
    "VN": "under_negotiation",

    # Framework agreed
    "US": "framework_agreed",
}

# Sector breakdown (global, from BAZG Annual Report 2024)
SECTORS = {
    "exports": [
        {"sector": "Chemical & pharma",       "sector_code": "CHEM_PHARMA",   "share_pct": 49.4, "value_CHF_millions": 139753},
        {"sector": "Machines & electronics",  "sector_code": "MACHINES_ELEC", "share_pct": 12.0, "value_CHF_millions":  33948},
        {"sector": "Watches & clocks",        "sector_code": "WATCHES",        "share_pct":  9.8, "value_CHF_millions":  27724},
        {"sector": "Precision & medtech",     "sector_code": "PRECISION",      "share_pct":  6.5, "value_CHF_millions":  18389},
        {"sector": "Metals",                  "sector_code": "METALS",         "share_pct":  5.3, "value_CHF_millions":  14994},
        {"sector": "Precious metals & gems",  "sector_code": "PRECIOUS",       "share_pct":  4.5, "value_CHF_millions":  12731},
        {"sector": "Textiles & clothing",     "sector_code": "TEXTILES",       "share_pct":  2.1, "value_CHF_millions":   5941},
        {"sector": "Agriculture & forestry",  "sector_code": "AGRI",           "share_pct":  2.0, "value_CHF_millions":   5658},
        {"sector": "Vehicles",               "sector_code": "VEHICLES",       "share_pct":  1.8, "value_CHF_millions":   5092},
        {"sector": "Other",                  "sector_code": "OTHER",          "share_pct":  7.1, "value_CHF_millions":  20070},
    ],
    "imports": [
        {"sector": "Chemical & pharma",       "sector_code": "CHEM_PHARMA",   "share_pct": 30.7, "value_CHF_millions":  68246},
        {"sector": "Machines & electronics",  "sector_code": "MACHINES_ELEC", "share_pct": 15.6, "value_CHF_millions":  34679},
        {"sector": "Precious metals & gems",  "sector_code": "PRECIOUS",       "share_pct": 12.0, "value_CHF_millions":  26676},
        {"sector": "Vehicles",               "sector_code": "VEHICLES",       "share_pct":  9.5, "value_CHF_millions":  21119},
        {"sector": "Metals",                  "sector_code": "METALS",         "share_pct":  7.1, "value_CHF_millions":  15783},
        {"sector": "Energy",                 "sector_code": "ENERGY",         "share_pct":  5.8, "value_CHF_millions":  12893},
        {"sector": "Agriculture & forestry",  "sector_code": "AGRI",           "share_pct":  5.5, "value_CHF_millions":  12227},
        {"sector": "Textiles & clothing",     "sector_code": "TEXTILES",       "share_pct":  4.2, "value_CHF_millions":   9337},
        {"sector": "Precision & medtech",     "sector_code": "PRECISION",      "share_pct":  3.8, "value_CHF_millions":   8447},
        {"sector": "Other",                  "sector_code": "OTHER",          "share_pct":  5.8, "value_CHF_millions":  12893},
    ],
}

# Time series: annual totals 2015-2025 + monthly 2025-2026
TIMESERIES = {
    "annual": [
        {"year": 2015, "exports": 201800, "imports": 175400, "balance":  26400},
        {"year": 2016, "exports": 200400, "imports": 170200, "balance":  30200},
        {"year": 2017, "exports": 208800, "imports": 174500, "balance":  34300},
        {"year": 2018, "exports": 226600, "imports": 188500, "balance":  38100},
        {"year": 2019, "exports": 228400, "imports": 187200, "balance":  41200},
        {"year": 2020, "exports": 220100, "imports": 170100, "balance":  50000},
        {"year": 2021, "exports": 255800, "imports": 196200, "balance":  59600},
        {"year": 2022, "exports": 278600, "imports": 230200, "balance":  48400},
        {"year": 2023, "exports": 274000, "imports": 225000, "balance":  49000},
        {"year": 2024, "exports": 282900, "imports": 222300, "balance":  60600},
        {"year": 2025, "exports": 287000, "imports": 232700, "balance":  54300, "preliminary": True},
    ],
    "monthly_2025_2026": [
        {"period": "2025-01", "exports": 22800, "imports": 18200},
        {"period": "2025-02", "exports": 21900, "imports": 17900},
        {"period": "2025-03", "exports": 24100, "imports": 19500},
        {"period": "2025-04", "exports": 23400, "imports": 18800},
        {"period": "2025-05", "exports": 24300, "imports": 19600},
        {"period": "2025-06", "exports": 23700, "imports": 19100},
        {"period": "2025-07", "exports": 22900, "imports": 18600},
        {"period": "2025-08", "exports": 23100, "imports": 19000},
        {"period": "2025-09", "exports": 24500, "imports": 19800},
        {"period": "2025-10", "exports": 24800, "imports": 20100},
        {"period": "2025-11", "exports": 23600, "imports": 19400},
        {"period": "2025-12", "exports": 23900, "imports": 19700},
        {"period": "2026-01", "exports": 22100, "imports": 17400, "preliminary": True},
        {"period": "2026-02", "exports": 22142, "imports": 17818},
        {"period": "2026-03", "exports": 22355, "imports": 19612},
    ],
}


def download_xlsx(url: str) -> openpyxl.Workbook:
    print(f"  Downloading {url} ...", end=" ", flush=True)
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    wb = openpyxl.load_workbook(BytesIO(r.content))
    print(f"OK ({len(r.content)//1024}KB)")
    return wb


def parse_bazg_sheet(wb: openpyxl.Workbook) -> tuple[float, dict[str, float]]:
    """Return (total_CHF_millions, {country_name: CHF_millions}) from business-cycle columns (0-5)."""
    ws = wb[wb.sheetnames[0]]
    total = 0.0
    result: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        name = row[0]
        val = row[3]
        if not (name and isinstance(name, str) and isinstance(val, (int, float))):
            continue
        if name == "Trading partners":
            continue
        if name == "Total":
            total = float(val)
        else:
            result[name] = round(float(val), 2)
    return total, result


def normalize_name(name: str) -> str:
    return name.strip()


def build_partners(
    exports: dict[str, float],
    imports: dict[str, float],
) -> list[dict]:
    all_names = set(exports) | set(imports)
    partners = []
    unmapped: list[str] = []

    for name in sorted(all_names):
        exp = exports.get(name, 0.0)
        imp = imports.get(name, 0.0)
        if exp + imp < MIN_VOLUME_MN:
            continue

        iso2 = NAME_TO_ISO2.get(name)
        if not iso2:
            unmapped.append(name)
            continue

        if iso2 == "CH":
            continue  # skip self

        centroid = CENTROIDS.get(iso2)
        fta = FTA_STATUS.get(iso2, "none")

        partner = {
            "country": name,
            "country_code": iso2,
            "exports": round(exp, 1),
            "imports": round(imp, 1),
            "balance": round(exp - imp, 1),
            "fta_status": fta,
        }
        if centroid:
            partner["centroid"] = centroid

        partners.append(partner)

    if unmapped:
        print(f"\n  [warn] {len(unmapped)} countries not in NAME_TO_ISO2 (volume ≥ {MIN_VOLUME_MN}M):")
        for n in sorted(unmapped):
            print(f"    '{n}': exports={exports.get(n,0):.0f}, imports={imports.get(n,0):.0f}")

    # Sort by total trade volume descending
    partners.sort(key=lambda p: -(p["exports"] + p["imports"]))
    return partners


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print("=== Swiss trade data download ===\n")

    print("1. Downloading BAZG export data...")
    exp_wb = download_xlsx(BAZG_EXP_URL)
    print("2. Downloading BAZG import data...")
    imp_wb = download_xlsx(BAZG_IMP_URL)

    print("\n3. Parsing...")
    total_exp, exp_raw = parse_bazg_sheet(exp_wb)
    total_imp, imp_raw = parse_bazg_sheet(imp_wb)
    print(f"   Exports: {len(exp_raw)} countries")
    print(f"   Imports: {len(imp_raw)} countries")
    print(f"   Totals: exports {total_exp:,.0f}M  imports {total_imp:,.0f}M  balance {total_exp-total_imp:+,.0f}M")

    print("\n4. Building partner list...")
    partners = build_partners(exp_raw, imp_raw)
    with_centroid = [p for p in partners if "centroid" in p]
    print(f"   {len(partners)} partners ≥ {MIN_VOLUME_MN}M CHF")
    print(f"   {len(with_centroid)} with centroids (shown as arcs)")

    output = {
        "metadata": {
            "source": "Federal Office for Customs and Border Security (BAZG), Annual Report 2024",
            "reference_year": 2024,
            "currency": "CHF",
            "unit": "millions",
            "note": "Business cycle total — excludes precious metals, gems, works of art and antiques",
            "downloaded": "2026-05-31",
            "total_exports": round(total_exp, 1),
            "total_imports": round(total_imp, 1),
            "trade_balance": round(total_exp - total_imp, 1),
        },
        "partners": partners,
        "sectors": SECTORS,
        "timeseries": TIMESERIES,
    }

    OUT_FILE.write_text(json.dumps(output, ensure_ascii=False, indent=2))
    size_kb = OUT_FILE.stat().st_size // 1024
    print(f"\n✓ Written to {OUT_FILE} ({size_kb}KB)")
    print(f"  Top 10 partners by total trade:")
    for p in partners[:10]:
        arc = "●" if "centroid" in p else "○"
        print(f"  {arc} {p['country_code']} {p['country']:25s}  exp={p['exports']:8,.0f}  imp={p['imports']:8,.0f}  bal={p['balance']:+9,.0f}  {p['fta_status']}")


if __name__ == "__main__":
    main()
