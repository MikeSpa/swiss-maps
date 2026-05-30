"""
Downloads the swisstopo commune typology (agglomeration category ACAT) and
outputs per-municipality typology data as public/demographics/typology.json
and public/demographics/typology_meta.json.

Source: swisstopo Boundaries 2022 — GEM_ORT_*_ANALYSE_22.xlsx inside the
boundaries ZIP (same file used by download_boundaries.py).
URL: https://dam-api.bfs.admin.ch/hub/api/dam/assets/21224783/master

Sheet g1a22 columns used:
  GMDNR  — BFS commune number (= bfs_nummer in the GeoJSON)
  ACAT   — agglomeration category (1–6; not all communes appear)

ACAT → typology class mapping:
  1 (core city)              → 1 "urban"
  2 (suburban belt)          → 1 "urban"
  3 (secondary city)         → 1 "urban"
  4 (periurban)              → 2 "periurban"
  6 (isolated urban centre)  → 2 "periurban"
  not in sheet               → 3 "rural"

Only ~1418 communes (those in/near agglomerations) appear in the sheet;
the remaining ~730 are assigned class 3 (rural).

Usage:
  uv run python scripts/download_typology.py
  uv run python scripts/download_typology.py --force   # re-download ZIP
"""

import argparse
import io
import json
import zipfile
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).parent.parent.parent
OUT_DIR = ROOT / "public" / "demographics"
CACHE_ZIP = OUT_DIR / "_swisstopo_boundaries.zip"
GEO_MUNICIPALITIES = ROOT / "public" / "geo" / "municipalities.geojson"

ZIP_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/21224783/master"

# ACAT value → typology class
ACAT_TO_CLASS: dict[int, int] = {
    1: 1,  # core city          → urban
    2: 1,  # suburban belt      → urban
    3: 1,  # secondary city     → urban
    4: 2,  # periurban          → periurban
    6: 2,  # isolated urban ctr → periurban
    # 5 is not defined in the research; treat as rural (fallback)
}

TYPOLOGY_META: dict[str, dict[str, str]] = {
    "1": {
        "de": "Städtisch",
        "fr": "Urbain",
        "it": "Urbano",
        "rm": "Urban",
        "en": "Urban",
    },
    "2": {
        "de": "Periurban",
        "fr": "Périurbain",
        "it": "Periurbano",
        "rm": "Periurban",
        "en": "Periurban",
    },
    "3": {
        "de": "Ländlich",
        "fr": "Rural",
        "it": "Rurale",
        "rm": "Rural",
        "en": "Rural",
    },
}


# ── Download ───────────────────────────────────────────────────────────────────

def download_zip(force: bool = False) -> Path:
    if CACHE_ZIP.exists() and not force:
        print(f"  ZIP already cached ({CACHE_ZIP.name}) — use --force to re-download")
        return CACHE_ZIP
    print(f"  Downloading {ZIP_URL}")
    resp = requests.get(ZIP_URL, timeout=180, headers={"User-Agent": "swiss-maps-pipeline/1.0"})
    resp.raise_for_status()
    CACHE_ZIP.write_bytes(resp.content)
    print(f"  Saved {CACHE_ZIP.name} ({CACHE_ZIP.stat().st_size / 1024:.0f} KB)")
    return CACHE_ZIP


# ── Parse ──────────────────────────────────────────────────────────────────────

def parse_acat(zip_path: Path) -> dict[int, int]:
    """Return {bfs_nummer: typology_class} by reading sheet g1a22 in the Excel."""
    with zipfile.ZipFile(zip_path) as zf:
        # Find the Excel file inside the ZIP (name varies by year)
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            raise RuntimeError(f"No .xlsx file found inside {zip_path.name}")
        # Prefer the file that contains "ANALYSE" in its name; fall back to first
        analyse = [n for n in xlsx_names if "ANALYSE" in n.upper()]
        xlsx_name = analyse[0] if analyse else xlsx_names[0]
        print(f"  Reading {xlsx_name}")
        xlsx_bytes = zf.read(xlsx_name)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    if "g1a22" not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise RuntimeError(f"Sheet 'g1a22' not found. Available: {available}")

    ws = wb["g1a22"]
    rows = iter(ws.iter_rows(values_only=True))

    # Find column indices from header row
    header = next(rows)
    col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    if "GMDNR" not in col or "ACAT" not in col:
        raise RuntimeError(f"Expected GMDNR and ACAT columns; found: {list(col)}")
    idx_bfs = col["GMDNR"]
    idx_acat = col["ACAT"]

    result: dict[int, int] = {}
    skipped = 0
    for row in rows:
        try:
            bfs = int(row[idx_bfs])
            acat = int(row[idx_acat])
        except (TypeError, ValueError):
            skipped += 1
            continue
        typology_class = ACAT_TO_CLASS.get(acat, 3)
        result[bfs] = typology_class

    print(f"  {len(result)} communes with explicit ACAT ({skipped} rows skipped)")
    return result


# ── Known communes ─────────────────────────────────────────────────────────────

def load_all_bfs_numbers() -> list[int]:
    """Return all bfs_nummer values from the municipalities GeoJSON (if available)."""
    if not GEO_MUNICIPALITIES.exists():
        print(f"  Note: {GEO_MUNICIPALITIES.name} not found — rural class assigned only to explicit absentees")
        return []
    geo = json.loads(GEO_MUNICIPALITIES.read_text())
    bfs_numbers = []
    for feat in geo["features"]:
        bfs = feat["properties"].get("bfs_nummer")
        if bfs is not None:
            bfs_numbers.append(int(bfs))
    print(f"  {len(bfs_numbers)} municipalities loaded from GeoJSON")
    return bfs_numbers


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Download swisstopo agglomeration typology for Swiss communes."
    )
    parser.add_argument("--force", action="store_true", help="Re-download the ZIP")
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Downloading swisstopo boundaries ZIP...")
    zip_path = download_zip(force=args.force)

    print("Parsing ACAT typology from sheet g1a22...")
    acat_map = parse_acat(zip_path)

    # Merge with full municipality list: anything not in the sheet → rural (3)
    all_bfs = load_all_bfs_numbers()
    typology: dict[str, int] = {}

    if all_bfs:
        for bfs in all_bfs:
            typology[str(bfs)] = acat_map.get(bfs, 3)
        rural_count = sum(1 for v in typology.values() if v == 3)
        urban_count = sum(1 for v in typology.values() if v == 1)
        periurban_count = sum(1 for v in typology.values() if v == 2)
        print(
            f"  Urban: {urban_count}, Periurban: {periurban_count}, Rural: {rural_count} "
            f"({len(typology)} total)"
        )
    else:
        # Without GeoJSON, output only what the sheet provides
        for bfs, cls in sorted(acat_map.items()):
            typology[str(bfs)] = cls
        print(f"  {len(typology)} communes with explicit typology (no GeoJSON to fill gaps)")

    # Write typology.json
    out_typology = OUT_DIR / "typology.json"
    out_typology.write_text(json.dumps(typology, ensure_ascii=False, separators=(",", ":")))
    print(f"  Written {out_typology.relative_to(ROOT)} ({out_typology.stat().st_size / 1024:.0f} KB)")

    # Write typology_meta.json
    out_meta = OUT_DIR / "typology_meta.json"
    out_meta.write_text(json.dumps(TYPOLOGY_META, ensure_ascii=False, indent=2))
    print(f"  Written {out_meta.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
